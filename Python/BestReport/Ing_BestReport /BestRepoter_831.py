import tkinter
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from tkinter.ttk import LabelFrame
from tkinter import filedialog

#######윈도우 생성#######
window=tkinter.Tk()
window.title("Best Reporter")
window.geometry("400x300+800+50")

window.resizable(False,False)

#######동작 함수#######
def fileOpen():
    #파일 탐색기 오픈해서 선택한 파일의 경로를 filename에 저장  
    filename=filedialog.askopenfilename(initialdir="",title="Select File")
    
    #######엑셀 변수 생성#######    row =1,2,3... // column=a,b,c.....
    newWb=Workbook()            #새로운 워크북 생성(우테 취합용)
    wb=load_workbook(filename)  #버튼으로 선택한 엑셀 파일 불러오기

    #버그리포트 취합 엑셀 전체 시트 중 뒤에서 1번째(우테 취합)까지 이름 가져오기    
    for ws,i in zip(wb.sheetnames, range(1,len(wb.sheetnames))): 
        brws = newWb.create_sheet(ws)
        count=2
        brws.cell(row=1,column=2).value="JIRA 링크"
        brws.cell(row=1,column=8).value="처리 여부"
        brws.cell(row=1,column=9).value="등급"
        brws.cell(row=1,column=10).value="등록자"
        brws.cell(row=1,column=12).value="글번호"
        brws.cell(row=1,column=13).value="담당자"

        for j in wb[ws].rows:        #해당 워크 시트 중 열의 끝까지 반복
            if (j[5].value == "JIRA 등록"): #해당 워크 시트의 5행의 열이 "JIRA 등록일 때"
                if(j[7].value == None):                         #JIRA 링크
                    brws.cell(row=count,column=2).value = "JIRA 링크 작성 누락"
                else:
                    brws.cell(row=count,column=2).value=j[7].value  
                brws.cell(row=count,column=8).value=j[5].value  #처리 여부
                brws.cell(row=count,column=9).value=j[6].value  #등급

                if(j[2].value == None):#등록자
                    brws.cell(row=count,column=10).value="이름없음"  #등록자
                else:
                    brws.cell(row=count,column=10).value=j[2].value  #등록자
                brws.cell(row=count,column=12).value=j[1].value  #글번호
                brws.cell(row=count,column=13).value=j[4].value  #담당자
                count=count+1
            elif (j[5].value =="본섭수정"):
                brws.cell(row=count,column=2).value=j[3].value  #글 요약
                brws.cell(row=count,column=8).value=j[5].value  #처리 여부
                brws.cell(row=count,column=9).value="B"  #등급
                if(j[2].value==None):
                    brws.cell(row=count,column=10).value="이름없음"  #등록자
                else:
                    brws.cell(row=count,column=10).value=j[2].value  #등록자
                brws.cell(row=count,column=12).value=j[1].value  #글번호
                brws.cell(row=count,column=13).value=j[4].value  #담당자
                count=count+1
            elif(j[5].value=="NextUpdate"):
                if(j[7].value == None):                         #JIRA 링크
                    brws.cell(row=count,column=2).value="JIRA 링크 작성 누락"    
                else:
                    brws.cell(row=count,column=2).value=j[7].value  
                brws.cell(row=count,column=8).value=j[5].value  #처리 여부
                brws.cell(row=count,column=9).value="B"  #등급
                if(j[2].value == None):
                    brws.cell(row=count,column=10).value="이름없음"  #등록자
                else:
                    brws.cell(row=count,column=10).value=j[2].value  #등록자
                brws.cell(row=count,column=12).value=j[1].value  #글번호
                brws.cell(row=count,column=13).value=j[4].value  #담당자
                count=count+1

    ###########c열에 vlookup 함수 입력해서 jira 값 이름 가져오기 ##############jira 이름 작성
    for ws, i in zip(newWb.sheetnames, range(1,len(newWb.sheetnames))):
        if (ws == "Sheet"):
            print(ws)
        else:
            fcount=2
            for r in newWb[ws].rows:               #워크시트의 열의 갯수만큼 반복
                ####함수 넣기#####
                if (newWb[ws].cell(row=fcount, column=8).value=="JIRA 등록"):     #워크시트의 처리 여부가 jira 등록일 경우 
                    newWb[ws].cell(row=fcount,column=3).value=f"=VLOOKUP(B{fcount},Sheet!B:C,2,0)"   # jira 제목 찾기

                elif (newWb[ws].cell(row=fcount, column=8).value=="NextUpdate"):  #워크시트의 처리 여부가 넥업일 경우
                    newWb[ws].cell(row=fcount,column=3).value=f"=VLOOKUP(B{fcount},Sheet!B:C,2,0)"   # jira 제목 찾기

                else:                   # 이외의 상황에서는 내용의 값을 찾아오기
                    newWb[ws].cell(row=fcount, column=3).value=newWb[ws].cell(row=fcount, column=2).value
                
                ####JIRA 링크 문자열 편집#####
                 #각 워크 시트의 B열의 값이 있다면
                if(newWb[ws].cell(row=fcount, column=2).value != None):    
                    jiraLink=newWb[ws].cell(row=fcount, column=2).value.strip() #공백 삭제
                    if (jiraLink.find("M") != -1):                      #문자열에 M이 있을 경우
                        jiraLink_index=jiraLink.index("M")                  #문자열 중 M이 있는 인덱스 값 찾기
                        jiraLink_Edit=jiraLink[jiraLink_index:jiraLink_index+10]    #JIRA 번호 추출

                        newWb[ws].cell(row=fcount, column=2).value=jiraLink_Edit      #셀에 작성해주기
                        fcount+=1
                    # 문자열에 M을 못 찾았거나 오류 발생 시
                    else:               
                        newWb[ws].cell(row=fcount, column=2).value = "Jira 링크 없음"
                        fcount+=1

    ########우테 카운팅하기#######
    #새로운 탭 만들고 등록자와 jira 등급 가져오기
    newWb.create_sheet("등록자 카운팅") #새로운 워크 시트 생성
    newWb['등록자 카운팅'].cell(row=1, column=2).value="JIRA 등급" # 제목 입력
    newWb['등록자 카운팅'].cell(row=1, column=3).value="등록자"    # 제목 입력
    total_count=2
    for fws,sc in zip(newWb.sheetnames, range(1,len(newWb.sheetnames))):  #1~워크시트 리스트 개수만큼 반복, fws = 개별 워크시크 이릅, sc=시트번호
        r_count=2
        if(sc >= 2):        #시트 넘버가 2번 이상일 경우
            for r in newWb[fws].rows:      #해당 워크시트의 열 개수만큼 반복
                if (newWb[fws].cell(row=r_count, column=9).value != None): # h행의 값이 빈 값이 아닐 경우 
                    newWb['등록자 카운팅'].cell(row=total_count, column=2).value = newWb[fws].cell(row=r_count, column=9).value  #새로운 워크 시트의 b행의 2열부터 값 입력(jira 등급)
                    reporterID=str(newWb[fws].cell(row=r_count, column=10).value).strip() #새로운 워크 시트의 c행의 2열부터 값 입력 (등록자)
                    newWb['등록자 카운팅'].cell(row=total_count, column=3).value = reporterID #공백 제거 후 값 입력
                    r_count+=1
                    total_count+=1

    ########등록자 중복 값 제거########
    newWb['등록자 카운팅'].cell(row=1, column=7).value="등록자 중복 제거"
    newWb['등록자 카운팅'].cell(row=1, column=8).value="B"
    newWb['등록자 카운팅'].cell(row=1, column=9).value="A"
    newWb['등록자 카운팅'].cell(row=1, column=10).value="S"
    newWb['등록자 카운팅'].cell(row=1, column=11).value="총 점"
    
    l_count=2
    l_count2=2
    b_count = 0
    a_count = 0
    s_count = 0
    reporterList = [] #등록자 리스트 저장할 변수
    reportGrade = []
    result = [] #등록자 리스트 중복이 제거된 값이 들어갈 변수
    for l in newWb['등록자 카운팅'].rows:
        reporterList.append(str(newWb['등록자 카운팅'].cell(row=l_count, column=3).value))
        reportGrade.append(str(newWb["등록자 카운팅"].cell(row=l_count, column=2).value))
        l_count+=1
        #중복값 제거 방법 
        for value in reporterList:  #reporterList 리스트를 value의 변수에 하나씩 저장
            if value not in result: #result에 value값이 없을 경우
                result.append(value)    #result 리스트에 값 저장
                newWb["등록자 카운팅"].cell(row=l_count2, column=7).value=value
                l_count2+=1

    dic_report={}
    for z in range(0, len(result)):
        for x in range(2,len(reporterList)+1):
            if (newWb['등록자 카운팅'].cell(row=x, column=3).value == result[z] and newWb["등록자 카운팅"].cell(row=x, column=2).value == "B"):
                b_count += 1
            elif (newWb['등록자 카운팅'].cell(row=x, column=3).value == result[z] and newWb["등록자 카운팅"].cell(row=x, column=2).value == "A"):
                a_count += 1
            elif (newWb['등록자 카운팅'].cell(row=x, column=3).value == result[z] and newWb["등록자 카운팅"].cell(row=x, column=2).value == "S"):
                s_count += 1
            else:
                pass 
        if(newWb['등록자 카운팅'].cell(row=z+2, column=7).value == "None"):
            newWb['등록자 카운팅'].cell(row=z+2, column=7).value = ""
        else:
            newWb["등록자 카운팅"].cell(row=z+2, column=8).value = b_count
            newWb["등록자 카운팅"].cell(row=z+2, column=9).value = a_count
            newWb["등록자 카운팅"].cell(row=z+2, column=10).value = s_count
            newWb["등록자 카운팅"].cell(row=z+2, column=11).value = b_count*1 + a_count*3 + s_count*10
            dic_report[newWb["등록자 카운팅"].cell(row=z+2, column=7).value] = int(newWb["등록자 카운팅"].cell(row=z+2, column=11).value)
        b_count = 0
        a_count = 0
        s_count = 0

    print(dic_report)
    newWb["등록자 카운팅"].cell(row=1,column=14).value = "랭킹 정렬"
    sort_dic_report = sorted(dic_report.items(), key=lambda item:item[1], reverse=True)
    for sort in range(0,len(sort_dic_report)):#14 등록자, 15점수
        newWb["등록자 카운팅"].cell(row=sort+2, column=14).value = sort_dic_report[sort][0]
        newWb["등록자 카운팅"].cell(row=sort+2, column=15).value = sort_dic_report[sort][1]
    newWb.save("BestTester.xlsx")
    messagebox.showinfo("BestTester.xlsx 파일 확인 부탁드립니다.")
    print("끝")

#######라벨 생성#######
#메인 라벨
mainLabelFrame=LabelFrame(window, text="우수 테스터 선발")
mainLabelFrame.place(relx=0.01,relwidth=0.98,relheight=0.99)

#서브 라벨1
subLabelFrame1=LabelFrame(mainLabelFrame, text="Setting")
subLabelFrame1.place(relx=0.01, relwidth=0.98, relheight=0.50)

#######버튼 생성#######
#파일 선택 버튼
fileopenBtn=tkinter.Button(subLabelFrame1,text="1) 유효 버그리포트 추출",width="30",command=fileOpen)
fileopenBtn.place(relx=0.1,rely=0.1)

window.mainloop()